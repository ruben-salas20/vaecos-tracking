"""Rutas del módulo financiero — Slices 2-3: listado + CRUD.

Slice 4 agregará analytics; Slice 5, catálogo admin.
"""
from __future__ import annotations

import re
from datetime import date

from flask import Blueprint, render_template, request, current_app, redirect, url_for, flash, session, abort

from ..auth.decorators import login_required, admin_required
from ..pagination import Pagination, read_pagination_args
from .repository import FinanzasRepository


_VALID_TIPOS = ("ingreso", "egreso", "transferencia")
_DATE_ISO_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


finanzas_bp = Blueprint("finanzas", __name__, url_prefix="/finanzas")


def _repo() -> FinanzasRepository:
    return FinanzasRepository(current_app.config["DB_PATH"])


_MONTHS_ES = [
    (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
    (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
    (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
]


def _parse_int(s):
    if s is None or s == "":
        return None
    try:
        return int(s)
    except (ValueError, TypeError):
        return None


@finanzas_bp.route("/")
@login_required
def list_movements():
    repo = _repo()

    years = repo.available_years()
    default_year = date.today().year
    if years and default_year not in years:
        default_year = years[0]

    raw_year = request.args.get("year")
    if raw_year is None:
        year = default_year
    else:
        parsed = _parse_int(raw_year)
        year = None if parsed == 0 else (parsed if parsed is not None else default_year)

    month = _parse_int(request.args.get("month"))
    if month == 0:
        month = None
    tipo = request.args.get("tipo") or None
    if tipo not in ("ingreso", "egreso", "transferencia"):
        tipo = None
    category_id = _parse_int(request.args.get("categoria"))
    if category_id == 0:
        category_id = None
    search = (request.args.get("q") or "").strip() or None

    page, per_page = read_pagination_args(default_per_page=50)

    movements, total = repo.list_movements(
        year=year, month=month, tipo=tipo, category_id=category_id,
        search=search, limit=per_page, offset=(page - 1) * per_page,
    )
    totals = repo.totals_for_filters(
        year=year, month=month, tipo=tipo, category_id=category_id, search=search,
    )
    pagination = Pagination.build(page=page, per_page=per_page, total=total)
    categories = repo.list_categories(only_active=False)

    return render_template(
        "finanzas/list.html",
        movements=movements,
        totals=totals,
        years=years,
        months=_MONTHS_ES,
        categories=categories,
        filters={
            "year": year if year is not None else 0,
            "month": month or 0,
            "tipo": tipo or "",
            "categoria": category_id or 0,
            "q": search or "",
        },
        pagination=pagination,
        default_year=default_year,
    )


# ── CRUD ──────────────────────────────────────────────────────────


def _parse_monto_to_centavos(raw: str) -> int | None:
    """'2.004.599,67' o '2004599.67' o '200000' → centavos. None si malformado."""
    if raw is None:
        return None
    s = raw.strip()
    if not s:
        return None
    # Detectar formato: si hay coma, asumimos formato colombiano (',' = decimal)
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    # Else: asumimos '.' como decimal (input HTML estándar)
    try:
        val = float(s)
    except ValueError:
        return None
    if val < 0:
        return None
    return int(round(val * 100))


def _parse_form_movement(form) -> tuple[dict | None, str | None]:
    """Devuelve (data, error_msg). Solo data válida o None."""
    fecha = (form.get("fecha") or "").strip()
    if not _DATE_ISO_RE.match(fecha):
        return None, "Fecha inválida — formato esperado YYYY-MM-DD."
    tipo = (form.get("tipo") or "").strip()
    if tipo not in _VALID_TIPOS:
        return None, "Tipo inválido."
    monto_centavos = _parse_monto_to_centavos(form.get("monto") or "")
    if monto_centavos is None or monto_centavos <= 0:
        return None, "Monto inválido (debe ser un número positivo)."
    observacion = (form.get("observacion") or "").strip()
    if not observacion:
        return None, "La observación es obligatoria."
    if len(observacion) > 500:
        return None, "La observación no puede exceder 500 caracteres."
    category_ids = [
        int(c) for c in form.getlist("categorias")
        if c.isdigit() and int(c) > 0
    ]
    if not category_ids:
        return None, "Seleccioná al menos una categoría."
    guia_ref = (form.get("guia_ref") or "").strip() or None
    return {
        "fecha": fecha,
        "tipo": tipo,
        "monto_centavos": monto_centavos,
        "observacion": observacion,
        "category_ids": category_ids,
        "guia_ref": guia_ref,
    }, None


@finanzas_bp.route("/new", methods=["GET"])
@login_required
def new_form():
    repo = _repo()
    categories = repo.list_categories(only_active=True)
    today = date.today().strftime("%Y-%m-%d")
    return render_template(
        "finanzas/form.html",
        movement=None,
        categories=categories,
        defaults={"fecha": today, "tipo": "egreso"},
    )


@finanzas_bp.route("/new", methods=["POST"])
@login_required
def new_save():
    data, err = _parse_form_movement(request.form)
    if err:
        flash(err, "error")
        return redirect(url_for("finanzas.new_form"))
    repo = _repo()
    actor = session.get("user_email") or "unknown"
    mov_id = repo.create_movement(
        fecha=data["fecha"], tipo=data["tipo"], monto_centavos=data["monto_centavos"],
        observacion=data["observacion"], category_ids=data["category_ids"],
        guia_ref=data["guia_ref"], creado_por=actor,
    )
    flash(f"Movimiento #{mov_id} creado correctamente.", "success")
    return redirect(url_for("finanzas.list_movements"))


def _can_edit(movement, user_email: str, role: str) -> bool:
    if role == "admin":
        return True
    return movement.creado_por == user_email


@finanzas_bp.route("/<int:mov_id>/edit", methods=["GET"])
@login_required
def edit_form(mov_id: int):
    repo = _repo()
    mov = repo.get_movement(mov_id)
    if not mov:
        abort(404)
    if not _can_edit(mov, session.get("user_email", ""), session.get("role", "user")):
        flash("No tenés permiso para editar este movimiento.", "error")
        return redirect(url_for("finanzas.list_movements"))
    categories = repo.list_categories(only_active=False)
    return render_template(
        "finanzas/form.html",
        movement=mov,
        categories=categories,
        defaults=None,
    )


@finanzas_bp.route("/<int:mov_id>/edit", methods=["POST"])
@login_required
def edit_save(mov_id: int):
    repo = _repo()
    mov = repo.get_movement(mov_id)
    if not mov:
        abort(404)
    if not _can_edit(mov, session.get("user_email", ""), session.get("role", "user")):
        flash("No tenés permiso para editar este movimiento.", "error")
        return redirect(url_for("finanzas.list_movements"))
    data, err = _parse_form_movement(request.form)
    if err:
        flash(err, "error")
        return redirect(url_for("finanzas.edit_form", mov_id=mov_id))
    actor = session.get("user_email") or "unknown"
    ok = repo.update_movement(
        mov_id,
        fecha=data["fecha"], tipo=data["tipo"], monto_centavos=data["monto_centavos"],
        observacion=data["observacion"], category_ids=data["category_ids"],
        guia_ref=data["guia_ref"], actualizado_por=actor,
    )
    if not ok:
        flash("No se pudo actualizar — el movimiento ya no existe.", "error")
    else:
        flash(f"Movimiento #{mov_id} actualizado.", "success")
    return redirect(url_for("finanzas.list_movements"))


@finanzas_bp.route("/<int:mov_id>/delete", methods=["POST"])
@admin_required
def delete(mov_id: int):
    repo = _repo()
    ok = repo.delete_movement(mov_id)
    if ok:
        flash(f"Movimiento #{mov_id} eliminado.", "success")
    else:
        flash(f"El movimiento #{mov_id} no existe.", "error")
    return redirect(url_for("finanzas.list_movements"))


# ── Analytics + Export ────────────────────────────────────────────


@finanzas_bp.route("/analytics")
@login_required
def analytics():
    repo = _repo()
    years = repo.available_years()
    default_year = date.today().year
    if years and default_year not in years:
        default_year = years[0]

    raw_year = request.args.get("year")
    if raw_year is None:
        year = default_year
    else:
        parsed = _parse_int(raw_year)
        year = None if parsed == 0 else (parsed if parsed is not None else default_year)

    totals = repo.totals_for_filters(year=year)
    breakdown = repo.category_breakdown(year=year)
    evolution = repo.monthly_evolution(year=year)

    return render_template(
        "finanzas/analytics.html",
        totals=totals,
        breakdown=breakdown,
        evolution=evolution,
        years=years,
        filters={"year": year if year is not None else 0},
        default_year=default_year,
    )


@finanzas_bp.route("/export")
@login_required
def export_xlsx():
    """Exporta los movimientos filtrados a Excel.

    Acepta los mismos query params que el listado.
    """
    import io
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from flask import send_file

    repo = _repo()

    raw_year = request.args.get("year")
    if raw_year is None:
        year = None
    else:
        parsed = _parse_int(raw_year)
        year = None if parsed == 0 else parsed
    month = _parse_int(request.args.get("month"))
    if month == 0:
        month = None
    tipo = request.args.get("tipo") or None
    if tipo not in _VALID_TIPOS:
        tipo = None
    category_id = _parse_int(request.args.get("categoria"))
    if category_id == 0:
        category_id = None
    search = (request.args.get("q") or "").strip() or None

    rows = repo.export_all(
        year=year, month=month, tipo=tipo,
        category_id=category_id, search=search,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Metadatos
    meta_parts = ["VAECOS · Finanzas"]
    if year is not None:
        meta_parts.append(f"Año: {year}")
    if month is not None:
        meta_parts.append(f"Mes: {month:02d}")
    if tipo:
        meta_parts.append(f"Tipo: {tipo}")
    meta_parts.append(f"Filas: {len(rows)}")
    meta_parts.append(f"Exportada: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    meta_cell = ws.cell(row=1, column=1, value=" · ".join(meta_parts))
    meta_cell.font = Font(italic=True, color="666666", size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    headers = ["ID", "Fecha", "Tipo", "Monto (COP)", "Observación", "Categorías", "Guía ref", "Creado por"]
    header_row = 3
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 26

    for i, row in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=row["id"])
        ws.cell(row=i, column=2, value=row["fecha"])
        ws.cell(row=i, column=3, value=row["tipo"])
        c = ws.cell(row=i, column=4, value=row["monto"])
        c.number_format = '#,##0.00'
        ws.cell(row=i, column=5, value=row["observacion"]).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=i, column=6, value=row["categorias"])
        ws.cell(row=i, column=7, value=row["guia_ref"])
        ws.cell(row=i, column=8, value=row["creado_por"])

    widths = [6, 12, 14, 16, 40, 30, 18, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if rows:
        ws.auto_filter.ref = f"A{header_row}:H{header_row + len(rows)}"

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    fecha = datetime.now().strftime("%Y-%m-%d")
    suffix = f"-{year}" if year else ""
    filename = f"finanzas-vaecos{suffix}-{fecha}.xlsx"
    return send_file(
        buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename,
    )


# ── Catálogo de categorías (admin) ────────────────────────────────


_HEX_COLOR_RE = re.compile(r"^#[0-9a-fA-F]{6}$")


@finanzas_bp.route("/categorias")
@admin_required
def categorias_list():
    repo = _repo()
    categories = repo.list_categories(only_active=False)
    usage = repo.category_usage_counts()
    return render_template(
        "finanzas/categorias.html",
        categories=categories,
        usage=usage,
    )


@finanzas_bp.route("/categorias", methods=["POST"])
@admin_required
def categorias_create():
    nombre = (request.form.get("nombre") or "").strip()
    color = (request.form.get("color") or "").strip() or None
    if not nombre:
        flash("El nombre es obligatorio.", "error")
        return redirect(url_for("finanzas.categorias_list"))
    if color and not _HEX_COLOR_RE.match(color):
        flash("Color inválido — debe ser hex tipo #RRGGBB.", "error")
        return redirect(url_for("finanzas.categorias_list"))
    cat_id = _repo().create_category(nombre, color)
    if cat_id is None:
        flash(f"Ya existe una categoría con el nombre {nombre!r}.", "error")
    else:
        flash(f"Categoría {nombre!r} creada.", "success")
    return redirect(url_for("finanzas.categorias_list"))


@finanzas_bp.route("/categorias/<int:cat_id>/edit", methods=["POST"])
@admin_required
def categorias_edit(cat_id: int):
    nombre = (request.form.get("nombre") or "").strip()
    color = (request.form.get("color") or "").strip() or None
    if not nombre:
        flash("El nombre es obligatorio.", "error")
        return redirect(url_for("finanzas.categorias_list"))
    if color and not _HEX_COLOR_RE.match(color):
        flash("Color inválido.", "error")
        return redirect(url_for("finanzas.categorias_list"))
    ok = _repo().update_category(cat_id, nombre=nombre, color=color)
    if not ok:
        flash("No se pudo actualizar (nombre duplicado o categoría inexistente).", "error")
    else:
        flash(f"Categoría #{cat_id} actualizada.", "success")
    return redirect(url_for("finanzas.categorias_list"))


@finanzas_bp.route("/categorias/<int:cat_id>/toggle", methods=["POST"])
@admin_required
def categorias_toggle(cat_id: int):
    new_state = _repo().toggle_category(cat_id)
    if new_state is None:
        flash(f"La categoría #{cat_id} no existe.", "error")
    else:
        flash(f"Categoría #{cat_id} {'activada' if new_state else 'desactivada'}.", "success")
    return redirect(url_for("finanzas.categorias_list"))
