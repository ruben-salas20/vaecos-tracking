from __future__ import annotations
import io
import sys
from datetime import datetime
from pathlib import Path
from flask import (
    Blueprint, render_template, request, session,
    redirect, url_for, flash, current_app, make_response, jsonify, send_file
)
from ..auth.decorators import login_required
from ..utils import sanitize_csv_field
from .jobs import create_job, get_job, dispatch_run, create_sync_job, get_sync_job, dispatch_sync

# Inject v0.3 for DashboardRepository
_V03_ROOT = Path(__file__).resolve().parents[3] / "v0.3"
if str(_V03_ROOT) not in sys.path:
    sys.path.insert(0, str(_V03_ROOT))

from vaecos_v03.storage import DashboardRepository

runs_bp = Blueprint("runs", __name__)


def _repo() -> DashboardRepository:
    return DashboardRepository(current_app.config["DB_PATH"])


@runs_bp.route("/run/new", methods=["POST"])
@login_required
def run_new():
    guides_raw = request.form.get("guides", "").strip()
    guides = [g.strip() for g in guides_raw.replace("\n", ",").split(",") if g.strip()]
    mode = request.form.get("mode", "dry-run")
    save_raw_html = bool(request.form.get("save_raw_html"))
    all_active = not guides

    token = create_job()
    dispatch_run(
        token=token,
        db_path=current_app.config["DB_PATH"],
        guides=guides,
        all_active=all_active,
        dry_run=(mode != "apply"),
        save_raw_html=save_raw_html,
    )
    return redirect(url_for("runs.run_progress", token=token))


@runs_bp.route("/run/progress/<token>")
@login_required
def run_progress(token: str):
    job = get_job(token)
    return render_template("dashboard/run_progress.html", token=token, job=job)


@runs_bp.route("/runs/<int:run_id>/export/effi")
@login_required
def export_effi(run_id: int):
    """Exporta a Excel las guías de la corrida que requieren 'Gestionar con encargado'.

    Formato organizado: header negro con marca, columnas con ancho ajustado,
    info de la corrida en una fila de metadatos arriba del data.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    repo = _repo()
    run = repo.get_run(run_id)
    rows = repo.export_effi_rows(run_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Corrida {run_id}"

    # ── Bloque de metadatos (fila 1) ─────────────────────────────────
    meta_parts = [f"VAECOS · Corrida #{run_id}"]
    if run:
        started = run["started_at"] if run["started_at"] else ""
        mode = run["mode"] if run["mode"] else ""
        if started:
            meta_parts.append(f"Iniciada: {started[:19].replace('T', ' ')}")
        if mode:
            meta_parts.append(f"Modo: {mode}")
    meta_parts.append(f"Filas: {len(rows)}")
    meta_parts.append(f"Exportada: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    meta_cell = ws.cell(row=1, column=1, value=" · ".join(meta_parts))
    meta_cell.font = Font(italic=True, color="666666", size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # ── Headers (fila 3, dejando 1 fila vacía de aire) ──────────────
    headers = ["No. Guía", "Estado actual (Effi)", "Problema", "Notas operadora"]
    header_row = 3
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 26

    # ── Datos (fila 4 en adelante) ──────────────────────────────────
    for i, row in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=row["guia"] or "")
        ws.cell(row=i, column=2, value=row["estado_effi_actual"] or "")
        ws.cell(row=i, column=3, value=sanitize_csv_field(row["motivo"] or ""))
        ws.cell(row=i, column=4, value=row["notas_operador"] or "")
        # Alineación + wrap para "Problema" y "Notas" (textos largos)
        ws.cell(row=i, column=3).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=i, column=4).alignment = Alignment(vertical="top", wrap_text=True)

    # ── Anchos de columna ──────────────────────────────────────────
    widths = [16, 28, 52, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze pane: el header siempre visible al hacer scroll ─────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── AutoFilter sobre el bloque de datos (Excel filter dropdowns)
    if rows:
        ws.auto_filter.ref = f"A{header_row}:D{header_row + len(rows)}"

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    fecha = datetime.now().strftime("%Y-%m-%d")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"vaecos-corrida-{run_id}-effi-{fecha}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@runs_bp.route("/sync/notion", methods=["POST"])
@login_required
def sync_notion():
    """Trigger a background sync of the guides snapshot from Notion."""
    token = create_sync_job()
    dispatch_sync(token=token, db_path=current_app.config["DB_PATH"])
    return redirect(url_for("runs.sync_progress", token=token))


@runs_bp.route("/sync/progress/<token>")
@login_required
def sync_progress(token: str):
    job = get_sync_job(token)
    return render_template("dashboard/sync_progress.html", token=token, job=job)


@runs_bp.route("/sync/status/<token>")
@login_required
def sync_status(token: str):
    """JSON status for AJAX polling."""
    job = get_sync_job(token)
    if not job:
        return jsonify({"status": "unknown"}), 404
    return jsonify(job)


@runs_bp.route("/runs/<int:run_id>/results/<path:guia>/notas", methods=["POST"])
@login_required
def update_notas(run_id: int, guia: str):
    note = request.form.get("notas_operador", "")
    _repo().update_operator_note(run_id, guia, note)
    # AJAX clients get JSON; classic form posts get the redirect fallback.
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True, "guia": guia, "notas_operador": note})
    return redirect(url_for("dashboard.run_detail", run_id=run_id))
