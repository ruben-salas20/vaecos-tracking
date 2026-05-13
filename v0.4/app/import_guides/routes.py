from __future__ import annotations
from datetime import datetime
from io import BytesIO
from flask import (
    Blueprint, render_template, request, session,
    redirect, url_for, flash, current_app, send_file
)
from ..auth.decorators import login_required
from .parser import ExcelParser, ParseResult

import_bp = Blueprint("import_guides", __name__)


@import_bp.route("/import", methods=["GET"])
@login_required
def import_form():
    return render_template("import_guides/import.html")


@import_bp.route("/import/template")
@login_required
def import_template():
    """Genera y descarga un Excel plantilla con los encabezados y 2 filas de ejemplo."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guías"

    headers = [
        "No. Guía",         # requerido
        "Cliente",          # opcional
        "Transportista",    # opcional (default: effi)
        "DPI",              # opcional (sus dígitos → teléfono)
        "Estado",           # opcional
        "Valor",            # opcional (numérico)
        "Contenido",        # opcional ("Nx PRODUCTO")
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    examples = [
        ["B263437621-1", "Ana Sofía Martínez",       "effi",      "12345678", "PENDIENTE",   179.00, "1 * CREMA ESTRECHANTE"],
        ["B263499000-1", "María José Hernández Pérez","effi",      "45782134", "EN RUTA",     358.00, "2 * CREMA ESTRECHANTE"],
        ["B263499001-1", "Carlos Eduardo Ramírez",    "effi",      "51234567", "",            76.00,  "1 * DERMAN"],
    ]
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    column_widths = [18, 32, 16, 14, 18, 12, 36]
    for i, w in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Hoja secundaria con notas
    notas_ws = wb.create_sheet("Notas")
    notas_lines = [
        "VAECOS · Plantilla de importación de guías",
        "",
        "Columnas reconocidas (encabezados en español, mayúsculas/minúsculas no importan):",
        "",
        "  - No. Guía: REQUERIDO. Número único de la guía. También acepta: 'tracking', 'guia',",
        "    'nro guía', 'guia transportadora'.",
        "  - Cliente: nombre completo del destinatario. Acepta: 'nombre', 'destinatario'.",
        "  - Transportista: 'effi' por default si se deja vacío. Acepta: 'carrier', 'empresa'.",
        "  - DPI: identificación del cliente. El sistema extrae los dígitos como teléfono.",
        "    Acepta: 'identificación', 'cédula', 'NIT', 'id destinatario'.",
        "  - Estado: estado inicial de la guía. Acepta: 'estado guía', 'estado actual'.",
        "  - Valor: monto en quetzales o moneda local (numérico). Acepta: 'monto', 'total'.",
        "  - Contenido: descripción del producto en formato 'N * PRODUCTO' (ej: '2 * CREMA",
        "    ESTRECHANTE'). El sistema separa cantidad y producto automáticamente.",
        "",
        "Reglas:",
        "  - Solo No. Guía es obligatorio. Si está vacío, la fila se descarta.",
        "  - Guías que ya existen en el sistema se marcan como 'omitidas' (no se duplican).",
        "  - Las primeras 3 filas se escanean buscando los encabezados — la fila exacta",
        "    de los headers no importa.",
        "",
        "Tras cargar el archivo, vas a ver una vista previa con: nuevas / omitidas / errores.",
        "Solo cuando confirmes, las guías nuevas se crean en Notion y se sincronizan a local.",
    ]
    for i, line in enumerate(notas_lines, start=1):
        cell = notas_ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
    notas_ws.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="vaecos-plantilla-importar-guias.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@import_bp.route("/import", methods=["POST"])
@login_required
def import_preview():
    from vaecos_v02.storage.db import connect as v02_connect
    db_path = current_app.config["DB_PATH"]
    f = request.files.get("excel_file")
    if not f or not f.filename:
        flash("Seleccioná un archivo Excel.", "error")
        return redirect(url_for("import_guides.import_form"))
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        flash("El archivo debe ser .xlsx o .xls.", "error")
        return redirect(url_for("import_guides.import_form"))

    conn = v02_connect(db_path)
    try:
        rows = conn.execute("SELECT DISTINCT guia FROM run_results").fetchall()
        existing = {r["guia"].upper() for r in rows}
    finally:
        conn.close()

    parser = ExcelParser(existing)
    result = parser.parse(f.stream, filename=f.filename)

    # Serialize preview to session (JSON-safe)
    session["import_preview"] = {
        "filename": result.filename,
        "raw_count": result.raw_count,
        "new": [
            {"guia": g.guia, "cliente": g.cliente, "carrier": g.carrier,
             "id_cliente": g.id_cliente, "telefono": g.telefono,
             "estado_inicial": g.estado_inicial, "valor": g.valor,
             "cantidad": g.cantidad, "producto": g.producto,
             "row_number": g.row_number}
            for g in result.new
        ],
        "skipped": [
            {"guia": g.guia, "cliente": g.cliente, "carrier": g.carrier,
             "id_cliente": g.id_cliente, "telefono": g.telefono,
             "estado_inicial": g.estado_inicial, "valor": g.valor,
             "cantidad": g.cantidad, "producto": g.producto,
             "row_number": g.row_number}
            for g in result.skipped
        ],
        "errors": [
            {"guia": g.guia, "error": g.error, "row_number": g.row_number}
            for g in result.errors
        ],
    }
    return render_template("import_guides/import_preview.html", result=result)


@import_bp.route("/import/confirm", methods=["POST"])
@login_required
def import_confirm():
    from vaecos_v02.storage.db import connect as v02_connect
    from vaecos_v02.providers.notion_provider import NotionProvider

    preview = session.pop("import_preview", None)
    if not preview:
        flash("No hay importación pendiente de confirmar.", "error")
        return redirect(url_for("import_guides.import_form"))

    settings = current_app.config["SETTINGS"]
    if not settings.notion_api_key or not settings.notion_data_source_id:
        flash("Faltan credenciales de Notion en .env (NOTION_API_KEY o NOTION_DATA_SOURCE_ID).", "error")
        return redirect(url_for("import_guides.import_form"))

    notion = NotionProvider(
        api_key=settings.notion_api_key,
        notion_version=settings.notion_version,
        data_source_id=settings.notion_data_source_id,
        query_kind="auto",
    )

    created: list[dict] = []
    failed: list[dict] = []
    for guide in preview["new"]:
        try:
            page_id = notion.create_guide_page(
                guia=guide["guia"],
                cliente=guide.get("cliente", ""),
                carrier=guide.get("carrier", "effi"),
                estado_novedad=guide.get("estado_inicial", ""),
                telefono=guide.get("telefono", ""),
                valor=guide.get("valor", ""),
                cantidad=guide.get("cantidad", 0) or 0,
                producto=guide.get("producto", ""),
            )
            created.append({**guide, "page_id": page_id})
        except Exception as exc:
            failed.append({**guide, "error": str(exc)})

    db_path = current_app.config["DB_PATH"]
    conn = v02_connect(db_path)
    try:
        conn.execute(
            "INSERT INTO import_log (imported_at, imported_by, filename, guides_new, guides_skipped, guides_error) "
            "VALUES (?,?,?,?,?,?)",
            (
                datetime.now().isoformat(timespec="seconds"),
                session.get("user_email", "unknown"),
                preview["filename"],
                len(created),
                len(preview["skipped"]),
                len(preview["errors"]) + len(failed),
            ),
        )
        conn.commit()
    finally:
        conn.close()

    return render_template(
        "import_guides/import_result.html",
        filename=preview["filename"],
        created=created,
        failed=failed,
        skipped_count=len(preview["skipped"]),
        parse_errors_count=len(preview["errors"]),
    )
